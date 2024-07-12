
import time
import math
import csv
import os
import locale
locale.setlocale(locale.LC_ALL, '')
import pandas as pd
from datetime import datetime
from csv import DictReader
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
import base64
from selenium.common.exceptions import TimeoutException
import re
import chardet

wb = load_workbook(filename="Automation.xlsx")
sheet = wb.active

xlsx_data = []
for value in sheet.iter_rows(values_only=True):
    xlsx_data.append(list(value))

with open("new.csv", "w") as csv_obj:
    writer = csv.writer(csv_obj, delimiter=",")
    for line in xlsx_data:
        writer.writerow(line)

# Detect the encoding of the CSV file
with open("new.csv", "rb") as f:
    result = chardet.detect(f.read())
encoding = result['encoding']

# Read the CSV file with the detected encoding
with open("new.csv", "r", encoding=encoding) as f:
    dict_reader = DictReader(f)
    csv_data = list(dict_reader)
os.remove("new.csv")



class Tags:
    sleepInSecond = 2
    EnterpriseID=f"//input[@spellcheck='true']"
    Continue=f"//button[@class='button accept']"
    LoginFirst=f"//button[normalize-space()='Login']"
    UserCode=f"//input[@id='usercode']"
    Password=f"//input[@id='password']"
    LoginSecond=f"//button[@type='submit']"
    Dropdown=f"//div[@class='combo-box has-focus']//div[@class='drop-btn']"
    DataBase=f"//span[normalize-space()='BUSIN16_PROD']"
    Continue_Accept=f"//button[@class='button accept']"
    LocateIcon=f"//div[@title='Locate (F2)']//i[@class='icon-Locate']"
    LogOut=f"//i[@class='icon-Logout']"
    LogoutSure=f"//body//root//message-box//button[1]"
    LocateButton=f"//button[@class='button accept']"
    HomeIcon=f"//span[contains(text(),'Home')]"
    BusinessName=f"(//asi-string-edit/div/input)[1]"
    EditIcon=f"//i[@class='icon-Edit']"
    GridData=f"(//div[@class='top-pad'])[6]/div"
    Policies=f"//span[normalize-space()='Policies']"
    PoliciesDropdown=f"//span[@class='frame-title-text context-menu-link']"
    AllExceptMarketed=f"//span[normalize-space()='All Except Marketed']"
    PoliciesGridData=f"(//div[@class='body-rows has-header']/div)[2]/div"
    AddIcon=f"//div[@title='Add']//i[@class='icon-Add']"
    Type=f"(//asi-combo-box/div/input)[2]"
    PolicyTextBox="(//asi-string-edit/div/input)[2]"
    EffectiveDate=f"(//asi-date-edit/div/input)[1]"
    ExpirationDate=f"(//asi-date-edit/div/input)[2]"
    Department=f"(//asi-combo-box/div/input)[16]"
    Premium=f"(//asi-currency-edit/div/input)[3]"
    Status=f"(//asi-combo-box/div/input)[8]"
    State=f"//div[@class='sidebar']//div[@class='body']//div[1]"
    IssuingLocation=f"(//asi-combo-box/div/input)[6]"
    LoginPopup=f"//div[@class='message-box']"
    YesBtn=f"//body//root//message-box//button[1]"
    IssuingCompany=f"(//asi-combo-box/div/input)[3]"
    PremiumPayable=f"(//asi-combo-box/div/input)[5]"
    FinishButton=f"//div[@class='proxy']//button[@class='button']"
    FinishButtonNew=f"//div[@data-automation-id='btnFinish']//button[@class='button accept']"
    ClosedRadioBtn=f"//div[@data-automation-id='rbtnClosed']//input[@name='pnlOpenCloseStatus']"
    YesButtonNew=f"//body//root//message-box//button[1]"
    RadioButton=f"//asi-radio-button/div/label/div"
    # testCancel=f"(//div[@class='close-button'])[2]"
    # testyes=f"//body//root//message-box//button[1]"


driver=webdriver.Chrome()
driver.maximize_window()
driver.get("https://busin16.appliedepic.com/#/")


try:
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, Tags.EnterpriseID))).send_keys(csv_data[0]["EnterpriseID"])
    print("Enterprise ID Inserted from Excel")
except:
    print(f"Failed to pass Enterprise ID in X-PATH: {Tags.EnterpriseID}")
time.sleep(Tags.sleepInSecond)


try:
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, Tags.Continue))).click()
    print("Clicked on the continue button")
except:
    print(f"Failed to Clicked on the continue button in X-PATH: {Tags.Continue}")
time.sleep(5)


try:
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, Tags.LoginFirst))).click()
    print("Clicked on the first time Login Button first")
except:
    print(f"Failed to Clicked on the first time Login Button in X-PATH: {Tags.LoginFirst}")
time.sleep(3)


# Switch to the new tab or window if it appears
main_window = driver.current_window_handle
try:
    WebDriverWait(driver, 30).until(
        EC.number_of_windows_to_be(2))  # Wait for the new window to open
    new_window = [window for window in driver.window_handles if window != main_window][0]
    driver.switch_to.window(new_window)
    print("Switched to the new window")
except:
    print("Failed to switch to the new window")
    driver.quit()


try:
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, Tags.UserCode))).send_keys(csv_data[0]["UserName"])
    print("User Name Inserted from Excel")
except:
    print(f"Failed to User Name Inserted from Excel in X-PATH: {Tags.UserCode}")
time.sleep(Tags.sleepInSecond)

try:
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, Tags.Password))).send_keys(csv_data[0]["Password"])
    print("Password Inserted from Excel")
except:
    print(f"Failed to Password Inserted from Excel in X-PATH: {Tags.Password}")
time.sleep(Tags.sleepInSecond)

# Continue with the next steps in the new window
try:
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, Tags.LoginSecond))).click()
    print("Clicked on the Login button Second time")
except:
    print(f"Failed to Clicked on the Login button Second time in X-PATH: {Tags.LoginSecond}")
time.sleep(15)

# Add any additional steps needed after switching to the new window

# Close the new window and switch back to the main window if necessary
# driver.close()
driver.switch_to.window(main_window)

try:
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, Tags.Dropdown))).click()
    print("Clicked on the Dropdown Icon")
except:
    print(f"Failed to PClicked on the Dropdown Icon in X-PATH: {Tags.Dropdown}")
time.sleep(Tags.sleepInSecond)


try:
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, Tags.DataBase))).click()
    print("Select BUSIN 16 PROD from dropdown")
except:
    print(f"Failed to Select BUSIN 16 PROD from dropdown in X-PATH: {Tags.DataBase}")
time.sleep(Tags.sleepInSecond)


try:
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, Tags.Continue_Accept))).click()
    print("Clicked on the continue button second time")
except:
    print(f"Failed to Clicked on the continue button second time in X-PATH: {Tags.Continue_Accept}")
time.sleep(3)


if Tags.LoginPopup: 
    try:
        WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, Tags.YesBtn))).click()
        print("Clicked on the Yes button")
    except:
        print(f"Failed to Clicked on the Yes button in X-PATH: {Tags.YesBtn}")
    time.sleep(5)
else:
    print("Login Popup Button not visible")


try:
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, Tags.LocateIcon))).click()
    print("Clicked on the Locate Icon")
except:
    print(f"Failed to Clicked on the Locate Icon in X-PATH: {Tags.LocateIcon}")
time.sleep(5)



for idx, row in enumerate(csv_data, start=2):  # Assuming the first row is the header
    if row["AccountName"]:
        try:
            WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.XPATH, Tags.BusinessName))).send_keys(row["AccountName"])
            print("Enter Business Name in the Search Box")
        except Exception as e:
            print(f"Failed to Enter Business Name in the Search Box in X-PATH: {Tags.BusinessName} - {str(e)}")  
        time.sleep(Tags.sleepInSecond)

        try:
            WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.XPATH, Tags.LocateButton))).click()
            print("Clicked on the Locate Button")
        except Exception as e:
            print(f"Failed to Clicked on the Locate Button in X-PATH: {Tags.LocateButton} - {str(e)}")  
        time.sleep(5)

        try:
            a = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, Tags.GridData))).text
            print(a)
            result = "Name Found"
            print("Found Grid data and write in the excel")
        except:
            result = "Name Not Found"
            print(f"Failed to find data for in X-PATH: {Tags.GridData}")
        time.sleep(Tags.sleepInSecond)

        # Write the result back to the Excel file
        sheet.cell(row=idx, column=len(xlsx_data[0]), value=result)


        if result=="Name Found":
            try:
                WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, Tags.EditIcon))).click()
                print("Clicked on the Edit Icon")
            except:
                print(f"Failed to Clicked on the Locate Button in X-PATH: {Tags.EditIcon}")  
            time.sleep(7)

            try:
                WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.XPATH, Tags.Policies))).click()
                print("Clicked on Policies icon from Navbar")
            except:
                print(f"Failed to Clicked on Policies icon from Navbar in X-PATH: {Tags.Policies}")  
            time.sleep(10)  

            try:
                WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.XPATH, Tags.PoliciesDropdown))).click()
                print("Clicked on Policies Dropdown")
            except:
                print(f"Failed to Clicked on Policies Dropdown in X-PATH: {Tags.PoliciesDropdown}")  
            time.sleep(Tags.sleepInSecond)

            try:
                WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.XPATH, Tags.AllExceptMarketed))).click()
                print("Clicked on All Except Marketed from Dropdown")
            except:
                print(f"Failed to Clicked on All Except Marketed from Dropdown in X-PATH: {Tags.AllExceptMarketed}")  
            time.sleep(6)


            # Extract the grid data and print it for debugging
            try:
                grid_elements = driver.find_elements(By.XPATH, Tags.PoliciesGridData)
                data = [element.text for element in grid_elements]
                print(f"Grid Data Elements: {data}")
            except Exception as e:
                print(f"Failed to extract grid data elements: {str(e)}")

            extracted_dates = []
            try:
                for element in grid_elements:
                    policies_text = element.text
                    print(f"Policies Text: {policies_text}")  # Debug print to check the text content

                    # Use regex to find all dates in the text
                    dates = re.findall(r'\b\d{1,2}/\d{1,2}/\d{4}\b', policies_text)
                    print(f"Found dates: {dates}")  # Debug print to check the found dates

                    # Append only the first date to the list
                    if dates:
                        extracted_dates.append(dates[0])

                print(f"Extracted Dates: {extracted_dates}")  # Print the list to verify the contents

            except Exception as e:
                print(f"Failed to print policies grid data in X-PATH: {Tags.PoliciesGridData} - {str(e)}")
            time.sleep(3)


            # Improved condition check
            record_found = False
            for element in data:
                if row["LOBCode"] in element and row["PolicyNumber"] in element:
                    for date in extracted_dates:
                        if row["Effective"] == date:
                            record_found = True
                            break 
                    if record_found:
                        break

            if record_found:
                print("Row data available in the grid section")
                record="Record Already Exit"
            else:
                print("Data Not found, creating a new record")
                record="Created New Record"

                try:
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.AddIcon))).click()
                    print("Clicked on the Add Icon")
                except Exception as e:
                    print(f"Failed to Clicked on the Add Icon in X-PATH: {Tags.AddIcon} - {str(e)}")
                time.sleep(8)


                try:
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.Type))).send_keys(row["LOBCode"])
                    print("Get LOB from Excel and insert value in the Type text box")
                except Exception as e:
                    print(f"Failed to Get LOB from Excel and insert value in the Type text box in X-PATH: {Tags.Type} - {str(e)}")
                time.sleep(Tags.sleepInSecond)


                try:
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.PolicyTextBox))).send_keys(row["PolicyNumber"])
                    print("Get Policy Number from Excel and insert value in the Policy text box")
                except Exception as e:
                    print(f"Failed to Get Policy Number from Excel and insert value in the Policy text box : {Tags.PolicyTextBox}")
                time.sleep(5)

                try:
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.EffectiveDate))).send_keys(row["Effective"])
                    print("Get Effective Date from Excel and insert value in the Effective text box")
                except Exception as e:
                    print(f"Failed to Get Effective Date from Excel and insert value in the Effective text box : {Tags.EffectiveDate}")
                time.sleep(Tags.sleepInSecond)


                try:
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.EffectiveDate))).send_keys(Keys.TAB)
                    print("Clicked Tab for Effective Date")
                except Exception as e:
                    print(f"Failed to Clicked Tab for Effective Date: {Tags.EffectiveDate}")
                time.sleep(4)


                try:
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.ExpirationDate))).clear()
                    print("Clear Expiration Date")
                except Exception as e:
                    print(f"Failed to Clear Expiration Date: {Tags.ExpirationDate}")
                time.sleep(Tags.sleepInSecond)


                try:
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.ExpirationDate))).send_keys(row["Expiration "])
                    print("Get Expiration  Date from Excel and insert value in the Expiration text box")
                except Exception as e:
                    print(f"Failed to Get Expiration Date from Excel and insert value in the Expiration text box {Tags.ExpirationDate}")
                time.sleep(3)


                try:
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.Department))).send_keys(row["Department"])
                    print("Get Department data from Excel and insert value in the Department text box")
                except Exception as e:
                    print(f"Failed to Get Department data from Excel and insert value in the Department text box: {Tags.Department}")
                time.sleep(Tags.sleepInSecond)


                try:
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.Premium))).send_keys(row["Premium"])
                    print("Get Premium data from Excel and insert value in the Premium text box")
                except Exception as e:
                    print(f"Failed to Get Premium data from Excel and insert value in the Premium text box: {Tags.Premium}")
                time.sleep(5)


                try:
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.Status))).send_keys(row["TransactionTypeCode"])
                    print("Get Transaction Type Code data from Excel and insert value in the Status text box")
                except Exception as e:
                    print(f"Failed to Get Transaction Type Code data from Excel and insert value in the Status text box: {Tags.Status}")
                time.sleep(3)


                try:
                    address = WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.State))).text
                    print("Print Address Data", address)
                    # Extract the state abbreviation assuming it follows the city and a comma
                    state_match = re.search(r'\b[A-Z][a-zA-Z]*,\s*([A-Z]{2})\b', address)
                    if state_match:
                        state = state_match.group(1)
                        print("Extracted State:", state)
                    else:
                        print("State not found in the address")
                except Exception as e:
                    print(f"Failed to print Address {Tags.State}: {e}")
                time.sleep(3)


                try:
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.IssuingLocation))).send_keys(state)
                    print("Enter State code in the Issuing Location Dropdown")
                except Exception as e:
                    print(f"Failed to Enter State code in the Issuing Location Dropdown {Tags.IssuingLocation}")
                time.sleep(3)



                try:
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.IssuingCompany))).send_keys(row["BillingCompanyCode"])
                    print("Enter Billing Company Code in the Issuing Company Dropdown")
                except Exception as e:
                    print(f"Failed to Enter Billing Company Code in the Issuing Company Dropdown {Tags.IssuingCompany}")
                time.sleep(3)


                try:
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.PremiumPayable))).clear()
                    print("Clear Premium Payable Text Box")
                except Exception as e:
                    print(f"Failed to Clear Premium Payable Text Box {Tags.PremiumPayable}")
                time.sleep(3)


                try:
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.PremiumPayable))).send_keys(row["MasterCompanyCode"])
                    print("Enter Master Company code in the Premium Payable Text Box")
                except Exception as e:
                    print(f"Failed to Enter Master Company code in the Premium Payable Text Box {Tags.PremiumPayable}")
                time.sleep(3)


                radio_button_value = row.get("RadioButton", "")
                if radio_button_value:
                    try:
                        radio_buttons = driver.find_elements(By.XPATH, Tags.RadioButton)
                        for button in radio_buttons:
                            if button.text.strip().lower() == radio_button_value.strip().lower():
                                button.click()
                                print(f"Clicked on the radio button '{radio_button_value}' for row {idx}")
                                break
                    except:
                        print(f"Failed to click on the radio button '{radio_button_value}' for row {idx}")
                else:
                    print(f"No radio button value provided for row {idx}")


                actions = ActionChains(driver)

                # Perform scroll by sending PAGE_DOWN key several times
                for _ in range(5):  # Adjust the range as needed
                    actions.send_keys(Keys.PAGE_DOWN).perform()
                    time.sleep(1)  # Adjust the sleep time as needed


                # try:
                #     WebDriverWait(driver, 30).until(
                #         EC.element_to_be_clickable((By.XPATH, Tags.testCancel))).click()
                #     print("clicked on the test cancel icon")
                # except Exception as e:
                #     print(f"clicked on the test cancel icon {Tags.testCancel}")
                # time.sleep(3)


                # try:
                #     WebDriverWait(driver, 30).until(
                #         EC.element_to_be_clickable((By.XPATH, Tags.testyes))).click()
                #     print("clicked on the test cancel icon")
                # except Exception as e:
                #     print(f"clicked on the test cancel icon {Tags.testyes}")
                # time.sleep(5)


                try:
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.FinishButton))).click()
                    print("Clicked On the Finish Button")
                except Exception as e:
                    print(f"Failed to Clicked On the Finish Button {Tags.FinishButton}")
                time.sleep(3)


                try:
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.ClosedRadioBtn))).click()
                    print("Clicked On the Closed Radio Button")
                except Exception as e:
                    print(f"Failed to Clicked On the Closed Radio Button {Tags.ClosedRadioBtn}")
                time.sleep(3)


                try:
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.FinishButtonNew))).click()
                    print("Clicked On the Finish Button Second Time")
                except Exception as e:
                    print(f"Failed to Clicked On the Finish Button Second Time {Tags.FinishButtonNew}")
                time.sleep(3)


                try:
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, Tags.YesButtonNew))).click()
                    print("Clicked On the Yes Button")
                except Exception as e:
                    print(f"Failed to Clicked On the Yes Button {Tags.YesButtonNew}")
                time.sleep(12)


            sheet.cell(row=idx, column=len(xlsx_data[0])+1, value=record)
            xlsx_data.append(row)  # Append the modified row with status


            try:
                WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((By.XPATH, Tags.LocateIcon))).click()
                print("Clicked on the Locate Icon")
            except Exception as e:
                print(f"Failed to Clicked on the Locate Icon in X-PATH: {Tags.LocateIcon} - {str(e)}")  
            time.sleep(Tags.sleepInSecond)

            try:
                business_name_element = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((By.XPATH, Tags.BusinessName)))
                business_name_element.clear()
                print("Clear Business Name from search box")
            except Exception as e:
                print(f"Failed to Clear Business Name from search box in X-PATH: {Tags.BusinessName} - {str(e)}")  
            time.sleep(Tags.sleepInSecond)


        else:
            try:
                WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((By.XPATH, Tags.LocateIcon))).click()
                print("Clicked on the Locate Icon")
            except Exception as e:
                print(f"Failed to Clicked on the Locate Icon in X-PATH: {Tags.LocateIcon} - {str(e)}")  
            time.sleep(Tags.sleepInSecond)

            try:
                business_name_element = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((By.XPATH, Tags.BusinessName)))
                business_name_element.clear()
                print("Clear Business Name from search box")
            except Exception as e:
                print(f"Failed to Clear Business Name from search box in X-PATH: {Tags.BusinessName} - {str(e)}")  
            time.sleep(Tags.sleepInSecond)

        wb.save("Automation.xlsx")
